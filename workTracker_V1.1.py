# -*- coding: utf-8 -*-
"""
Created on Sat Jul  8 12:40:49 2023

@author: hnewey7
"""

import os
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import ttk
import datetime

# Initialisation.
def init():
    
    print("Starting Initialisation...")
    
    # Check if spreadsheet exists.
    if checkForSpreadsheet()==False:
        
        print("No spreadsheet.")
        
        # Open new spreadsheet.
        workbook = Workbook()
        
        # Get active sheet.
        sheet = workbook.active
        
        # Setting headings
        sheet['A1'] = 'Date:'
        sheet['B1'] = 'Start Time:'
        sheet['C1'] = 'End Time:'
        sheet['D1'] = 'Tasks:'
        
        # Setting column widths and wrapping text.
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['D'].width = 50
        
        # Save workbook as specific name.
        workbook.save("Work Tracker.xlsx")
        
        print("Spreadsheet created.") 
        
    print("Launching GUI...")
        
    # Launch GUI
    GUI()
        
# Check for spreadsheet.   
def checkForSpreadsheet():
    
    # Get folder path.
    folderPath = os.path.dirname(os.path.abspath(__file__))
    
    # File path for spreadsheet.
    filePath = folderPath + "\\Work Tracker.xlsx"
    
    # Check if file exists.
    if os.path.exists(filePath):
        return True
    else:
        return False

# Launch GUI
def GUI():
    
    # Function to update time.
    def updateTime():
        
        # Get current time.
        currentTime = datetime.datetime.now()
        time = str(currentTime).split(".")
        displayTime = time[0]
        
        # Creating timer label element.
        timer = tk.Label(window, text=displayTime, pady=20)
        
        # Formatting timer label elements.
        timer.grid(row=1,column=1,columnspan=4)
        
        # Loop this every second.
        window.after(1000, updateTime)
        
        return()
        
    # Function to set new date on spreadspreed.
    def updateDate():
        
        # Get current date.
        currentTime = datetime.datetime.now()
        date = str(currentTime).split(" ")
        displayDate = date[0]
        
        # Read spreadsheet and sheet.
        workbook = load_workbook("Work Tracker.xlsx")
        sheet = workbook.active
        
        # Get the number of rows.
        rowNumber = sheet.max_row
        
        # If dates have already been entered.
        if rowNumber > 1:
        
            # Getting value for the last cell.
            lastCell = sheet['A' + str(rowNumber)].value
        
            # If previous date not equal to current date, and the day is not the weekend.
            if lastCell!=displayDate:
                
                # Add current date.
                sheet['A' + str(rowNumber + 1)] = displayDate
            
        else:
            
            # Enter the first date.
            sheet['A2'] = displayDate
        
        # Save workbook as specific name.
        workbook.save("Work Tracker.xlsx")
        
        # Loop this every minute.
        window.after(60000, updateDate)
    
    # Function to set the start time.
    def setStartTime():
        print("Start time set.")
        
        # Read spreadsheet and sheet.
        workbook = load_workbook("Work Tracker.xlsx")
        sheet = workbook.active
        
        # Get the number of rows.
        rowNumber = sheet.max_row
        
        # Get start time.
        hours = startHours.get()
        mins = startMins.get()
        formattedHours = str(hours).zfill(2)
        formattedMins = str(mins).zfill(2)
        time = formattedHours + ":" + formattedMins
        
        # Set value in spreadsheet.
        sheet['B' + str(rowNumber)].value = time
        
        # Save workbook as specific name.
        workbook.save("Work Tracker.xlsx")
    
    # Functon to set the end time.
    def setEndTime():
        print("End time set.")
        
        # Read spreadsheet and sheet.
        workbook = load_workbook("Work Tracker.xlsx")
        sheet = workbook.active
        
        # Get the number of rows.
        rowNumber = sheet.max_row
        
        # Get start time.
        hours = endHours.get()
        mins = endMins.get()
        formattedHours = str(hours).zfill(2)
        formattedMins = str(mins).zfill(2)
        time = formattedHours + ":" + formattedMins
        
        # Set value in spreadsheet.
        sheet['C' + str(rowNumber)].value = time
        
        # Save workbook as specific name.
        workbook.save("Work Tracker.xlsx")
        
    def addTask():
        print("Task Added.")
        
        # Get task data.
        additionalTask = task.get()
        
        # Read spreadsheet and sheet.
        workbook = load_workbook("Work Tracker.xlsx")
        sheet = workbook.active
        
        # Get the number of rows.
        rowNumber = sheet.max_row
        
        # Get previous task data.
        tasks = sheet['D' + str(rowNumber)].value
        print(tasks)
        
        if tasks!=None:
            
            # Update task.
            tasks = str(tasks) + " " + additionalTask
        
        else:
            
            # First new task.
            tasks = additionalTask
            
        # Set value in spreadsheet.
        sheet['D' + str(rowNumber)].value = tasks
        
        # Save workbook as specific name.
        workbook.save("Work Tracker.xlsx")
        
        # Clearing task entry.
        addTaskText.delete(0, tk.END)
    
    # Creating window with tkinter.
    window = tk.Tk()
    window.title("Work Tracker")
    window.geometry("700x300")
    
    # Creating variables.
    startHours = tk.IntVar()
    startMins = tk.IntVar()
    endHours = tk.IntVar()
    endMins = tk.IntVar()
    task = tk.StringVar()
    
    # Creating start time elements.
    startTimeHours = ttk.Spinbox(window, width=3, from_=0, to=23, increment=1, textvariable=startHours)
    startTimeMins = ttk.Spinbox(window, width=3, from_=0, to=55, increment=5, textvariable=startMins)
    startTimeButton = tk.Button(window, text="Set Start Time", command=setStartTime, width=15)
    
    # Creating end time elements.
    endTimeHours = ttk.Spinbox(window, width=3, from_=0, to=23, increment=1, textvariable=endHours)
    endTimeMins = ttk.Spinbox(window, width=3, from_=0, to=55, increment=5, textvariable=endMins)
    endTimeButton = tk.Button(window, text="Set End Time", command=setEndTime, width=15)
    
    # Creating add task elements.
    addTaskText = tk.Entry(window, width=30, textvariable=task)
    addTaskButton = tk.Button(window, text="Add Task", command=addTask, width=10)
    
    # Formatting start time elements.
    startTimeHours.grid(row=2,column=1, pady=10)
    startTimeMins.grid(row=2,column=2, padx=5)
    startTimeButton.grid(row=2,column=3, padx=10)
    
    # Formatting end time elements.
    endTimeHours.grid(row=3,column=1)
    endTimeMins.grid(row=3,column=2)
    endTimeButton.grid(row=3,column=3)
    
    # Formatting add task elements.
    addTaskText.grid(row=2,column=4, padx=30)
    addTaskButton.grid(row=3,column=4)
    
    # Configuring rows to align vertically.
    window.grid_rowconfigure(0, weight=30)
    window.grid_rowconfigure(1, weight=1)
    window.grid_rowconfigure(2, weight=1)
    window.grid_rowconfigure(3, weight=1)
    window.grid_rowconfigure(4, weight=100)
    
    # Configuring columns to align horizontally.
    window.grid_columnconfigure(0, weight=100)
    window.grid_columnconfigure(1, weight=1)
    window.grid_columnconfigure(2, weight=1)
    window.grid_columnconfigure(3, weight=1)
    window.grid_columnconfigure(4, weight=1)
    window.grid_columnconfigure(5, weight=100)
    
    # Updating time and label.
    updateTime()
    
    # Updating date for spreadsheet.
    updateDate()
    
    # Lifting window 
    window.lift()
    
    # Starting main loop of GUI.
    window.mainloop()

init()


